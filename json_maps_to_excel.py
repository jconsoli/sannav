#!/usr/bin/python
# -*- coding: utf-8 -*-
# Copyright 2023, 2026 Jack Consoli.  All rights reserved.
#
# NOT BROADCOM SUPPORTED
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may also obtain a copy of the License at
# https://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
"""
:mod:`json_maps_to_excel` - Converts a SANnav MAPS export, which is a json dump, to Excel.

Version Control::

    +-----------+---------------+-----------------------------------------------------------------------------------+
    | Version   | Last Edit     | Description                                                                       |
    +===========+===============+===================================================================================+
    | 4.0.0     | 04 Aug 2023   | Re-Launch                                                                         |
    +-----------+---------------+-----------------------------------------------------------------------------------+
    | 4.0.1     | 20 Feb 2026   | Updated copyright notice.                                                         |
    +-----------+---------------+-----------------------------------------------------------------------------------+
"""

__author__ = 'Jack Consoli'
__copyright__ = 'Copyright 2023, 2026 Jack Consoli'
__date__ = '20 Feb 2026'
__license__ = 'Apache License, Version 2.0'
__email__ = 'jack_consoli@yahoo.com'
__maintainer__ = 'Jack Consoli'
__status__ = 'Released'
__version__ = '4.0.1'

import collections
import argparse
import os
import copy
import openpyxl.utils.cell as xl
import brcdapi.log as brcdapi_log
import brcdapi.gen_util as gen_util
import brcdapi.excel_util as excel_util
import brcdapi.excel_fonts as excel_fonts
import brcdapi.file as brcdapi_file
import brcddb.brcddb_common as brcddb_common
import brcddb.report.utils as report_utils

_DOC_STRING = False  # Should always be False. Prohibits any code execution. Only useful for building documentation
_DEBUG = False
_DEBUG_i = 'rene/mod_moderate_policy_48'
_DEBUG_o = None
_DEBUG_log = '_logs'
_DEBUG_nl = None

_link_font = excel_fonts.font_type('link')
_hdr_1_font = excel_fonts.font_type('hdr_1')
_hdr_2_font = excel_fonts.font_type('hdr_2')
_bold_font = excel_fonts.font_type('bold')
_std_font = excel_fonts.font_type('std')
_align_wrap = excel_fonts.align_type('wrap')
_align_wrap_c = excel_fonts.align_type('wrap_center')
_align_wrap_vc = excel_fonts.align_type('wrap_vert_center')

_tc_headers = collections.OrderedDict()  # d: Report header, c: column width, v: print header vertical
_tc_headers['policyName'] = dict(d='Policy', c=32, ha='wrap', da='wrap')
_tc_headers['virtualSwitchId'] = dict(d='FID', c=5, ha='wrap', da='wrap')
_tc_headers['tags'] = dict(d='Tags', c=32, ha='wrap', da='wrap')
_tc_headers['description'] = dict(d='Description', c=32, ha='wrap', da='wrap')
_tc_headers['isActive'] = dict(d='Active', c=5, ha='wrap_vert_center', da='wrap_center')
_tc_headers['isCustom'] = dict(d='Custom', c=5, ha='wrap_vert_center', da='wrap_center')

def _place_holder(v):
    """Place holder for methods intended to convert something in a rule.

    :param v: ruleType value
    :type v: str, int, bool, float, dict, list, None
    :return: Converted value
    :rtype: int
    """
    return v


def _generic_list(v):
    """Converts a list to a string.

    :param v: ruleType value
    :type v: str, int, bool, float, dict, list, None
    :return: Converted value
    :rtype: int
    """
    return ', '.join(v) if isinstance(v, list) else 'NOT LIST (' + str(type(v)) + ')'


"""Rule & measure conversions
d     User friendly description. If not present, assume there are sub-dictionaries to process
c     Column width
v     Boolean. If True, print header vertically
m     Method to call to convert the value
ha    Header alignment
da    Display of values alignment
o     Apparently, you can't have nested ordered dictionaries so this table is used to order the output
k     Dictionary key. Added in _descriptors()
col   Column number. Added in _create_sheet()"""
_rule_conv = dict(
    o=('groupName', 'groupType', 'isDefaultGroup', 'ruleName', 'ruleType', 'isDefaultRule', 'baseRuleName',
       'severityType', 'quietTime', 'measureDetails'),
    groupName=dict(d='Group Name', c=32, ha=_align_wrap, da=_align_wrap),
    groupType=dict(d='Group Type', c=5, m=_place_holder, ha=_align_wrap_vc, da=_align_wrap_c),
    isDefaultGroup=dict(d='Default Group', c=5, ha=_align_wrap_vc, da=_align_wrap_c),
    ruleName=dict(d='Rule Name', c=38, ha=_align_wrap, da=_align_wrap),
    ruleType=dict(d='Rule Type', c=5, m=_place_holder, ha=_align_wrap_vc, da=_align_wrap_c),
    isDefaultRule=dict(d='Default Rule', c=5, ha=_align_wrap_vc, da=_align_wrap_c),
    baseRuleName=dict(d='Base Rule Name', c=32, ha=_align_wrap, da=_align_wrap),
    severityType=dict(d='Severity Type', c=5, m=_place_holder, ha=_align_wrap_vc, da=_align_wrap_c),
    quietTime=dict(d='Quiet Time', c=8, ha=_align_wrap, da=_align_wrap),
    measureDetails=dict(
        o=('measureId', 'thresholdDtls', 'timeBaseDtls', 'swActions'),
        measureId=dict(d='Measure', c=24, ha=_align_wrap, da=_align_wrap),
        thresholdDtls=dict(
            o=('operator', 'thresholdList', 'thresholdValue'),
            operator=dict(d='Operator', c=5, m=_place_holder, ha=_align_wrap_vc, da=_align_wrap_c),
            thresholdList=dict(d='Threshold List', c=12, m=_generic_list, ha=_align_wrap, da=_align_wrap),
            thresholdValue=dict(d='Threshold Value', c=18, ha=_align_wrap_c, da=_align_wrap_c)
        ),
        timeBaseDtls=dict(
            o=('timeBaseValue',),
            timeBaseValue=dict(d='Time Base Value', c=5, ha=_align_wrap_vc, da=_align_wrap_c)
        ),
        swActions=dict(
            o=('rasLogEvent', 'snmpTrap', 'email', 'portDecommission', 'fence', 'sfpStatusMarginal', 'fms', 'sddq',
               'unQuarantine', 'toggle', 'switchStatusCritical', 'switchStatusMarginal', 'reBalance', 'fpin'),
            rasLogEvent=dict(d='RAS Log', c=5, ha=_align_wrap_vc, da=_align_wrap_c),
            snmpTrap=dict(d='SNMP Trap', c=5, ha=_align_wrap_vc, da=_align_wrap_c),
            email=dict(d='email', c=5, ha=_align_wrap_vc, da=_align_wrap_c),
            portDecommission=dict(d='Decommission', c=5, ha=_align_wrap_vc, da=_align_wrap_c),
            fence=dict(d='Fence', c=5, ha=_align_wrap_vc, da=_align_wrap_c),
            sfpStatusMarginal=dict(d='SFP Marginal', c=5, ha=_align_wrap_vc, da=_align_wrap_c),
            fms=dict(d='FMS', c=5, ha=_align_wrap_vc, da=_align_wrap_c),
            sddq=dict(d='SDDQ', c=5, ha=_align_wrap_vc, da=_align_wrap_c),
            unQuarantine=dict(d='Un-quarantine', c=5, ha=_align_wrap_vc, da=_align_wrap_c),
            toggle=dict(d='Toggle', c=5, ha=_align_wrap_vc, da=_align_wrap_c),
            switchStatusCritical=dict(d='Switch Critical', c=5, ha=_align_wrap_vc, da=_align_wrap_c),
            switchStatusMarginal=dict(d='Switch Marginal', c=5, ha=_align_wrap_vc, da=_align_wrap_c),
            reBalance=dict(d='Re-balance', c=5, ha=_align_wrap_vc, da=_align_wrap_c),
            fpin=dict(d='FPIN', c=5, ha=_align_wrap_vc, da=_align_wrap_c),
        ),
    ),
)


def _create_sheet(wb, tc, sheet_name, sheet_i, sheet_title, headers):
    """Creates a worksheet for the Excel report.

    :param wb: Workbook object
    :type wb: class
    :param tc: Table of context page. A link to this page is place in cell A1
    :type tc: str, None
    :param sheet_name: Sheet (tab) name
    :type sheet_name: str
    :param sheet_i: Sheet index where page is to be placed.
    :type sheet_i: int
    :param sheet_title: Title to be displayed in large font, hdr_1, at the top of the sheet
    :type sheet_title: str
    :param headers: List of dictionaries: d: Header, c: column width
    :type headers: list, tuple
    :return: openxl sheet
    :rtype: Worksheet
    """
    global _link_font, _hdr_1_font, _hdr_2_font, _align_wrap_vc, _align_wrap
    
    # Create the worksheet
    sheet = wb.create_sheet(index=sheet_i, title=sheet_name)
    sheet.page_setup.paperSize = sheet.PAPERSIZE_LETTER
    sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
    row = col = 1

    # Add the link back to the table of contents and title
    excel_util.cell_update(sheet, row, col, 'Contents', font=_link_font, link='#' + tc + '!A1')
    col += 1
    sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=len(headers)-1)
    excel_util.cell_update(sheet, row, col, sheet_title, font=_hdr_1_font)

    # Add the headers and set the column width
    row, col = row+2, 1
    sheet.freeze_panes = sheet['A' + str(row+1)]
    for d in headers:
        sheet.column_dimensions[xl.get_column_letter(col)].width = d['c']
        excel_util.cell_update(sheet, row, col, d['d'], font=_hdr_2_font, align=d['ha'])
        d.update(dict(col=col))
        col += 1

    return sheet


def _descriptors(obj, base_ref):
    """Recursively looks for 'd' in a dict and builds a list of descriptor dictionaries

    :param obj: Object to parse
    :type obj: dict
    :param base_ref: Base key reference
    :type base_ref: str
    :return: List of column widths
    :rtype: list
    """
    r_list = list()
    for k in obj['o']:
        d = obj.get(k)
        if isinstance(d, dict):
            if d.get('c') is not None:
                d0 = copy.deepcopy(d)
                d0.update(dict(k=base_ref + k))
                r_list.append(d0)
            else:
                r_list.extend(_descriptors(d, base_ref + k + '/'))

    return r_list


def _rule_value(d, rule):
    """Finds and returns the value for a matching key in / notation for a rule

    :param d: One of the dictionaries in _rule_conv
    :type d: dict
    :param rule: The rule from the SANnav MAPS export
    :type rule: dict
    :return: The value in the rule matching the key, 'k', in d. If 'm' is in d, then it is converted using that method
    :rtype: list, tuple, str, bool, int, float, None
    """
    if d.get('k') is None:
        return None

    kl = d.get('k').split('/')
    v0 = rule
    while len(kl) > 0:
        v0 = v0.get(kl.pop(0))
    return v0 if d.get('m') is None else d.get('m')(v0)


def _maps_page(wb, tc, sheet_name, sheet_i, sheet_title, rule_list):
    """Creates a worksheet for a rule group to the Excel report.

    :param wb: Workbook object
    :type wb: class
    :param tc: Table of context page. A link to this page is placed in cell A1
    :type tc: str, None
    :param sheet_name: Sheet (tab) name
    :type sheet_name: str
    :param sheet_i: Sheet index where page is to be placed.
    :type sheet_i: int
    :param sheet_title: Title to be displayed in large font, hdr_1, at the top of the sheet
    :type sheet_title: str
    :param rule_list: List of rules
    :type rule_list: list, tuple
    """
    global _rule_conv, _std_font, _align_wrap, _bold_font

    # Add the sheet to the workbook
    descriptor_list = _descriptors(_rule_conv, '')
    sheet = _create_sheet(wb, tc, sheet_name, sheet_i, sheet_title, descriptor_list)

    # Fill out the sheet
    row = 4
    for rule in rule_list:
        if isinstance(rule, str):
            excel_util.cell_update(sheet, row, 1, rule, font=_bold_font)
        else:
            for d in descriptor_list:
                buf = None
                v = _rule_value(d, rule)
                if v is not None:
                    if isinstance(v, bool):
                        buf = '\u221A' if v else ''
                    else:
                        buf = ', '.join(v) if isinstance(v, (list, tuple)) else v
                excel_util.cell_update(sheet, row, d['col'], buf, font=_std_font, align=d['da'])
        row += 1


_page = dict(  # t: Sheet Title, n: Sheet name, m: method to call to create sheet
    # Note: The method to call to build the Workbook sheet is specified with 'm'. I did it this way so that I could
    # define what action to take from a table. As it turned out, all rules returned from SANnav are consistent so 'm'
    # is always the pointer to the same method, _maps_page(). Having worked with the FOS API, I'm conditioned to
    # having to deal with many special cases.
    rulesUnderPortCategory=dict(t='Port Rules', n='port', m=_maps_page),
    rulesUnderSwitchStatusCategory=dict(t='Switch Status Rules', n='switch_status', m=_maps_page),
    rulesUnderFabricCategory=dict(t='Fabric Rules', n='fabric', m=_maps_page),
    rulesUnderFruCategory=dict(t='FRU Rules', n='FRU', m=_maps_page),
    rulesUnderSecurityCategory=dict(t='Security Rules', n='security', m=_maps_page),
    rulesUnderResourceCategory=dict(t='Resource Rules', n='resource', m=_maps_page),
    rulesUnderTrafficOrFlowsCategory=dict(t='Traffic and Flow Rules', n='traffic', m=_maps_page),
    rulesUnderFpiCategory=dict(t='Fabric Performance Impact Rules', n='fpi', m=_maps_page),
    rulesUnderExtensionTunnelCategory=dict(t='Extension Tunnel Rules', n='ext_tunnel', m=_maps_page),
    rulesUnderExtensionGePortCategory=dict(t='GE Port Rules', n='ext_ge_port', m=_maps_page),
    rulesUnderBackendPortCategory=dict(t='Backend Port Rules', n='backend', m=_maps_page),
    rulesUnderIoLatencyCategory=dict(t='I/O Latency Rules', n='latency', m=_maps_page),
    rulesUnderIoPerformanceCategory=dict(t='Performance Rules', n='performance', m=_maps_page),
    rulesUnderIoSCSICategory=dict(t='SCSI Rules', n='scsi', m=_maps_page),
    rulesUnderAmpHealthCategory=dict(t='AMP Health Rules', n='amp_health', m=_maps_page),
    rulesUnderAmpResourceCategory=dict(t='AMP Resource Rules', n='amp_resource', m=_maps_page),
    rulesUnderIoHealthCategory=dict(t='I/O Health Rules', n='io_health', m=_maps_page),
    uncategorizedDefaultRules=dict(t='Uncategorized Default Rules', n='uncat_dflt', m=_maps_page),
)


def _write_report(json_obj, report):
    """Creates an Excel workbook with the port statistics differences. Operates off global data

    :param json_obj: JSON MAPS object to be converted to a Workbook
    :type json_obj: dict
    :param report: Name of report (Excel file name)
    :type report: str
    :return: Status code. See brcddb.brcddb_common.EXIT_xxxx
    :rtype: int
    """
    global _page, _tc_headers

    # Create an Excel Workbook
    wb = excel_util.new_report()

    # Setup the data structure, t_content, for the table of contents.
    tc_page = 'MAPS_Policies'
    t_content = [dict(merge=6, font='hdr_1', align='wrap', disp='MAPS Policies and Rules Report'), dict()]
    max_cols = 0
    for v in _tc_headers.values():
        t_content.append(dict(new_row=False, font='hdr_2', align=v['ha'], disp=v['d']))
        max_cols += 1
    t_content.append(dict())

    # Set up the table of contents
    for obj in json_obj:
        sheet_index = 0

        # Add the sub-header
        for k, v in _tc_headers.items():
            buf = obj.get(k)
            if isinstance(buf, bool):
                buf = '\u221A' if buf else ''
            elif not isinstance(buf, str):
                buf = ''
            t_content.append(dict(new_row=False, font='hdr_2', align=v['da'], disp=buf))
        t_content.append(dict())

        # Create and add the sheets to the Workbook
        all_l = list()
        content = dict() if obj.get('categoryDetailsInfo') is None else obj.get('categoryDetailsInfo')
        for k1, v1 in content.items():
            t_content.append(dict(new_row=False, font='std', align='wrap', disp=''))
            try:
                sheet_name = _page[k1]['n'] + '_' + str(sheet_index)
                sheet_title = _page[k1]['t']
                action = _page[k1]['m']
            except KeyError:
                sheet_name = 'unknown_' + str(sheet_index)
                sheet_title = k1
                action = _maps_page
            t_content.append(dict(merge=max_cols-3, font='link', align='wrap', disp=sheet_title,
                                  hyper='#' + sheet_name + '!A1'))
            all_l.append(k1)
            sorted_l = gen_util.sort_obj_str(v1, ('groupName', 'ruleName'))
            all_l.extend(sorted_l)
            action(wb, tc_page, sheet_name, sheet_index, sheet_title, sorted_l)
            sheet_index += 1

        # Add an "All Rules" page
        sheet_name = 'all_' + str(sheet_index)
        sheet_title = 'All'
        t_content.append(dict(new_row=False, font='std', align='wrap', disp=''))
        t_content.append(dict(merge=max_cols-3, font='link', align='wrap', disp=sheet_title,
                              hyper='#' + sheet_name + '!A1'))
        _maps_page(wb, tc_page, sheet_name, sheet_index, sheet_title, all_l)

    # Insert the table of contents as the title page at the beginning of the Workbook and save the report
    report_utils.title_page(wb, None, tc_page, 0, 'MAPS Policies', t_content, [d['c'] for d in _tc_headers.values()])
    excel_util.save_report(wb, report)

    return brcddb_common.EXIT_STATUS_OK


def _get_input():
    """Parses the module load command line

    :return i: Input file name
    :rtype i: str
    :return r: Report name
    :rtype r: str
    """
    global _DEBUG, _DEBUG_i, _DEBUG_REPORT

    if _DEBUG:
        args_i, args_o, args_log, args_nl =\
            brcdapi_file.full_file_name(_DEBUG_i, '.json'), _DEBUG_o, _DEBUG_log, _DEBUG_nl

    else:
        buf = 'Create an Excel Workbook of MAPS policies from the JSON output from SANnav.'
        parser = argparse.ArgumentParser(description=buf)
        buf = '(Required) Name of JSON MAPS file to convert. The ".json" extension is automatically added.'
        parser.add_argument('-i', help=buf, required=True)
        buf = '(Optional) Name of Excel Workbook. The default is the input file name, with ".json" substituted for '\
              '".xlsx".'
        parser.add_argument('-o', help=buf, required=False)
        buf = '(Optional) Directory where log file is to be created. Default is to use the current directory. The ' \
              'log file name will always be "Log_xxxx" where xxxx is a time and date stamp.'
        parser.add_argument('-log', help=buf, required=False,)
        buf = '(Optional) No parameters. When set, a log file is not created. The default is to create a log file.'
        parser.add_argument('-nl', help=buf, action='store_true', required=False)
        args = parser.parse_args()
        args_i, args_o, args_log, args_nl = brcdapi_file.full_file_name(args.i, '.json'), args.o, args.log, args.nl

    # Setup the log file
    if not args_nl:
        brcdapi_log.open_log(args_log)

    # User feedback
    ml = [
        os.path.basename(__file__) + ' version: ' + __version__,
        'Input file:  ' + args_i,
        'Output file: ' + str(args_o),
        '',
    ]
    if _DEBUG:
        ml.insert(0, 'WARNING!!! Debug is enabled')
    brcdapi_log.log(ml, echo=True)

    return args_i, args_i.replace('.json', '.xlsx') if args_o is None else args_o


def pseudo_main():
    """Basically the main(). Did it this way so it can easily be used as a standalone module or called from another.

    :return: Exit code. See exit codes in brcddb.brcddb_common
    :rtype: int
    """
    global _DEBUG, __version__

    # Get and validate user input
    in_f, report = _get_input()

    # Read in the JSON MAPS file
    obj = brcdapi_file.read_dump(in_f)
    if obj is None:
        return brcddb_common.EXIT_STATUS_ERROR

    return _write_report(obj, report)


##################################################################
#
#                    Main Entry Point
#
###################################################################

if _DOC_STRING:
    print('_DOC_STRING is True. No processing')
else:
    brcdapi_log.close_log(str(pseudo_main()), True, True)
